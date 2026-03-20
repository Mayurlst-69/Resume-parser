import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from api.models import ParseJob, JobStatus

HEADER_FILL  = PatternFill("solid", fgColor="1D9E75")
WARN_FILL    = PatternFill("solid", fgColor="FAC775")
UNSURE_FILL  = PatternFill("solid", fgColor="FEF3C7")  # amber-100 for unsure cells
ALT_FILL     = PatternFill("solid", fgColor="F5FAF8")
NULL_FONT    = Font(italic=True, color="999999")
UNSURE_FONT  = Font(italic=True, color="B45309")       # amber for "?"
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)

thin   = Side(style="thin", color="E0E0E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

ALL_COLUMNS = [
    ("Name",         "name",         "extract_name"),
    ("Position",     "position",     "extract_position"),
    ("Phone",        "phone",        "extract_phone"),
    ("Email",        "email",        "extract_email"),
    ("Address",      "address",      "extract_address"),
    ("Education",    "education",    "extract_education"),
    ("Experience",   "experience",   "extract_experience"),
    ("Confidence",   "confidence",   None),   # always show
    ("Source File",  "filename",     None),
    ("Parse Method", "parse_method", None),
    ("Status",       "status",       None),
]
ALL_COL_WIDTHS = [25, 28, 18, 30, 35, 35, 40, 12, 38, 14, 12]

UNSURE_DISPLAY = "?"   # what Excel shows for unsure fields


def _cell_value(key: str, values: dict, result) -> tuple[str | None, str]:
    """
    Returns (display_value, cell_type)
    cell_type: "normal" | "null" | "unsure"
    """
    if key == "name":
        if result and result.name_cert == "unsure":
            return UNSURE_DISPLAY, "unsure"
        val = values.get("name")
        return (val, "normal") if val and val != "null" else (None, "null")

    if key == "position":
        if result and result.position_cert == "unsure":
            return UNSURE_DISPLAY, "unsure"
        val = values.get("position")
        return (val, "normal") if val and val != "null" else (None, "null")

    val = values.get(key)
    if val is None or val == "null" or val == "":
        return None, "null"
    return val, "normal"


def build_excel(jobs: list[ParseJob], config=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Parsed Resumes"

    # ── Filter columns based on config toggles ──
    active_cols = []
    active_widths = []
    for (label, key, config_key), width in zip(ALL_COLUMNS, ALL_COL_WIDTHS):
        if config_key is None:
            active_cols.append((label, key))
            active_widths.append(width)
        elif config and getattr(config, config_key, True):
            active_cols.append((label, key))
            active_widths.append(width)

    # Header
    for col_idx, (label, _) in enumerate(active_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        is_low_conf = job.status == JobStatus.low_confidence
        is_alt      = row_idx % 2 == 0
        r           = job.result

        values = {
            "name":         r.name if r else None,
            "position":     r.position if r else None,
            "phone":        r.phone if r else None,
            "email":        r.email if r else None,
            "address":      r.address if r else None,
            "education":    r.education if r else None,
            "experience":   r.experience if r else None,
            "confidence":   f"{int(r.confidence * 100)}%" if r else "—",
            "filename":     job.filename,
            "parse_method": job.parse_method or "—",
            "status":       job.status.value,
        }

        for col_idx, (_, key) in enumerate(active_cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            display, cell_type = _cell_value(key, values, r)

            if cell_type == "unsure":
                cell.value = UNSURE_DISPLAY
                cell.font  = UNSURE_FONT
                cell.fill  = UNSURE_FILL
            elif cell_type == "null":
                cell.value = "null"
                cell.font  = NULL_FONT
                if is_low_conf:
                    cell.fill = WARN_FILL
                elif is_alt:
                    cell.fill = ALT_FILL
            else:
                cell.value = display
                if is_low_conf:
                    cell.fill = WARN_FILL
                elif is_alt:
                    cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 18

    for col_idx, width in enumerate(active_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(active_cols))}1"

    # Summary sheet
    ws2 = wb.create_sheet(title="Summary")
    total   = len(jobs)
    done    = sum(1 for j in jobs if j.status == JobStatus.done)
    flagged = sum(1 for j in jobs if j.status == JobStatus.low_confidence)
    failed  = sum(1 for j in jobs if j.status == JobStatus.failed)
    unsure_count = sum(
        1 for j in jobs
        if j.result and (j.result.name_cert == "unsure" or j.result.position_cert == "unsure")
    )

    for r_idx, (label, val) in enumerate([
        ("Total files", total),
        ("Parsed OK", done),
        ("Low confidence (flagged)", flagged),
        ("Failed", failed),
        ("Fields marked '?' (needs review)", unsure_count),
    ], start=1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=val)

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()